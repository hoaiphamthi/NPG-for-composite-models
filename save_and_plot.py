import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook, Workbook
import openpyxl
import seaborn as sns
import matplotlib.patches as mpatches
import pickle
import pandas as pd

def save(data, name_instance,name, saving_obj):
    if saving_obj == 1:
        saved_obj = "data"
    elif saving_obj == 2:
        saved_obj = "results"
    else:
        print("Invalid saving_obj in save_result. Saving objective have to be either 1(for data) or 2(for result). Program exited.")
        exit()
    folder_path = os.path.join(saved_obj, name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    filename = os.path.join(folder_path, name_instance)
    np.save(filename, data)

def load(name_instance,name, saving_obj):
    if saving_obj == 1:
        saved_obj = "data"
    elif saving_obj == 2:
        saved_obj = "results"
    else:
        print("Invalid saving_obj in save_result. Saving objective have to be either 1(for data) or 2(for result). Program exited.")
        exit()
    folder_path = os.path.join(saved_obj, name)
    filename = os.path.join(folder_path, name_instance) + ".npy"
    loaded_data = np.load(filename, allow_pickle=True)
    return loaded_data.item()

def make_all_plots(results, name_instance,name,save = True, plot_mse = False, plot = True, legend = True):
    plot_res(results, "res", name_instance,name,save = save, plot=plot, legend=legend)
    plot_res(results, "obj", name_instance,name,save = save, plot=plot, legend=legend)
    plot_res(results, "obj", name_instance, name ,save = save,over_time=True, plot=plot, legend=legend)
    plot_res(results, "steps", name_instance, name,save = save, plot=plot, legend=legend)
    boxplot(results, name_instance, name,save = save, plot=plot)
    if plot_mse == True:
        plot_res(results, "mse", name_instance, name,save = save, plot=plot, legend=legend)


def find_min_value(results):
    f_min = np.inf
    for history_values in results.values():
        f_min = min(f_min, min(history_values["obj"]))
    return f_min

def assign_color(alg_names):
    color_map = {
    'red': [ "#fc8d59", '#d73027', "#92580b", "#045D2C" ],    # For 'NPG' group #fc8d59
    'green': ["#5dc404", "#045D2C", "#09EE70ED", "#022211EC"],             # For 'LS' group
    'blue': ["#6d36ab", '#2171b5', "#110388"]                # AdPG group
}
    assigned_colors = []

    # Track usage index for each group
    color_index = {'red': 0, 'green': 0, 'blue': 0}
    if not any("NPG1" in name for name in alg_names): color_map["red"] = color_map["red"][1:]
    for name in alg_names:
        if 'NPG' in name:
            group = 'red'
        elif 'LS' in name:
            group = 'green'
        else:
            group = 'blue'
        
        group_colors = color_map[group]
        index = color_index[group] % len(group_colors)
        assigned_colors.append(group_colors[index])
        color_index[group] += 1
    return assigned_colors

def assign_marker(alg_names):
    marker_map = {
    'NPG': ['o', '*', 'D', 'x'],    # For 'NPG' group #fc8d59
    'LS': ['<', '>', 'x'],             # For 'LS' group
    'AdPG': ['^', 'v','+']                # AdPG group
}
    assigned_markers = []

    # Track usage index for each group
    marker_index = {'NPG': 0, 'LS': 0, 'AdPG': 0}
    if not any("NPG1" in name for name in alg_names): marker_map["NPG"] = marker_map["NPG"][1:]
    for name in alg_names:
        if 'NPG' in name:
            group = 'NPG'
        elif 'LS' in name:
            group = 'LS'
        else:
            group = 'AdPG'
        
        # Pick next color in group's list (cycle if needed)
        group_markers = marker_map[group]
        index = marker_index[group] % len(group_markers)
        assigned_markers.append(group_markers[index])
        marker_index[group] += 1
    return assigned_markers

def plot_res(results, key, name_instance,name ,over_time = False, with_fopt = True, save = True, plot = True, legend = True):
    sns.set(style="whitegrid", context="paper")
    plt.figure()
    name_image = name_instance + "_"+ key
    fontsize = 18
    if over_time == True: 
        name_image += "_over_time"
        plt.xlabel('Time(s)', fontsize=fontsize)
    else:
        plt.xlabel('Iterations', fontsize=fontsize)
    if key == "obj":
        plt.ylabel('Objective', fontsize=fontsize)
    elif key == "res":
        plt.ylabel('Residual', fontsize=fontsize)
    elif key == "grad":
        plt.ylabel("Norm of gradient", fontsize=fontsize)
    elif key == "steps":
        plt.ylabel("Stepsize", fontsize=fontsize)
    elif key == "mse":
        plt.ylabel("Mean squared error", fontsize=fontsize)
    name_image += ".pdf"
    if key == "obj" and with_fopt:
        opt = find_min_value(results) 
    else:
        opt = 0.0
    i = 0
    numerical_stabilizer = 0 
    algs = results.keys()
    colors = assign_color(algs)
    markers = assign_marker(algs)
    for alg, history_values in results.items():
        N = len(history_values[key])
        values = np.array(history_values[key]) - opt + numerical_stabilizer
        if with_fopt: 
            mask = values > 0
        else:
            mask = [True for i in range(N)]
        printed_values = values[mask]
        if over_time == True:
            time_axis = np.array(history_values["time"])[mask]
            if len(set(time_axis)) != len(time_axis):
                print("Couldn't plot objective over time due to too small time step.")
                return None
            plt.plot(time_axis, printed_values, label = alg, marker = markers[i], markevery = int(N / 10) +1, color = colors[i], linewidth=1.5, markersize=7 ) 
        else:
            plt.plot(printed_values, label=alg, marker = markers[i], markevery = int(N / 10)+1, color = colors[i], linewidth=1.5, markersize=7)
        i += 1
    plt.grid(True)
    plt.tick_params(axis="both", which="major", labelsize=fontsize - 2, bottom=True, left=True)
    plt.yscale("log")
    plt.title(name_instance, fontsize=fontsize)
    if legend: plt.legend(fontsize=fontsize-5)
    plt.tight_layout()
    if save:
        folder_path = os.path.join("plots", name, name_instance)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        plt.savefig(os.path.join(folder_path, name_image), dpi=300)
    if plot:
        plt.show()
    plt.close()

def boxplot(results, name_instance, name , save = True, plot = True):
    name_image = name_instance + "_" + 'stepsize_boxplot'
    fontsize = 14
    name_image += ".pdf"
    
    data_values = list( d['steps'] for d in results.values())
    labels = list(results.keys())
    colors = assign_color(labels)  # Customize colors here

    # Create box plot
    sns.set(style="whitegrid", context="paper")
    fig, ax = plt.subplots()
    box = ax.boxplot(data_values, patch_artist=True, showfliers=False, medianprops=dict(color='black', linewidth=1.5))  # patch_artist=True enables fill color

    # Set colors
    for patch, color in zip(box['boxes'], colors):
        patch.set_facecolor(color)

    # Customize plot
    ax.set_xticks([])
    ax.set_ylabel('Stepsize', fontsize = fontsize)
    ax.set_xlabel('Algorithms', fontsize = fontsize)
    plt.tight_layout()
    handles = [mpatches.Patch(color=c, label=l) for c, l in zip(colors, labels)]
    ax.legend(handles=handles, loc='upper right', fontsize=fontsize - 4)
    if save:
        folder_path = os.path.join("plots", name, name_instance)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        plt.savefig(os.path.join(folder_path, name_image), dpi=300)
    if plot:
        plt.show()
    plt.close()

def plot_performance_profile(name, legend = True):
    with open(os.path.join("results", name + "_allresults.pkl"), "rb") as f:
        all_results = pickle.load(f)
    df_time = pd.DataFrame({
        size: {alg: all_results[size][alg]['time'] for alg in all_results[size]}
        for size in all_results
    })

    df_obj = pd.DataFrame({
        size: {alg: all_results[size][alg]['obj'] for alg in all_results[size]}
        for size in all_results
    })

    df_res = pd.DataFrame({
        size: {alg: all_results[size][alg]['res'] for alg in all_results[size]}
        for size in all_results
    })

    df_iter = pd.DataFrame({
        size: {alg: all_results[size][alg]['iteration'] for alg in all_results[size]}
        for size in all_results
    })

    df_steps = pd.DataFrame({
        size: {alg: all_results[size][alg]['steps'] for alg in all_results[size]}
        for size in all_results
    })

    
    def compute_performance_profile(df, numerical_stable = False, truncation = None):
            algs = df.index.tolist()
            matrix = np.array([np.concatenate(df.loc[a].values) for a in algs])
            if numerical_stable == True:
                matrix += 10**-20
            best_per_problem = np.min(matrix, axis=0)
            ratio = matrix / best_per_problem
            # The taus for plotting (sorted ratios)
            taus = np.unique(ratio.flatten())
            taus.sort()
            if truncation == None: truncation = np.min(np.max(ratio, axis=1))
            if numerical_stable == True:
                taus = taus[taus <= truncation]
            # Compute profiles
            profiles = {}
            P = matrix.shape[1]  # number of problems

            for i, alg in enumerate(algs):
                r = ratio[i, :]
                profiles[alg] = [np.sum(r <= t) / P for t in taus]
            return taus, profiles
    algs = df_iter.index.tolist()
    colors = assign_color(algs)
    markers = assign_marker(algs)

    taus, profiles = compute_performance_profile(df_obj, numerical_stable=True)
    plot_profile(taus, profiles, 'Performance profile for objective', name + '_obj', colors, markers, logscale=True, legend=legend)
    taus, profiles = compute_performance_profile(df_res, numerical_stable=True)
    plot_profile(taus, profiles, 'Performance profile for residual', name + '_res', colors, markers, logscale=True, legend=legend)
    taus, profiles = compute_performance_profile(df_iter)
    plot_profile(taus, profiles, 'Performance profile for iteration', name + '_iter', colors, markers, legend=legend)
    taus, profiles = compute_performance_profile(df_time)
    plot_profile(taus, profiles, 'Performance profile for running time', name + '_time', colors, markers, legend=legend)
    taus, profiles = compute_performance_profile(df_steps)
    plot_profile(taus, profiles, 'Performance profile for 1/stepsize', name + '_stepsize', colors, markers, legend=legend)

    if 'lasso' in name:
        df_mse = pd.DataFrame({
                size: {alg: all_results[size][alg]['mse'] for alg in all_results[size]}
                for size in all_results
            })
        taus, profiles = compute_performance_profile(df_mse, numerical_stable=True, truncation=1.0000005)
        plot_profile(taus, profiles, 'Performance profile for mean squared error', name + '_mse', colors, markers, legend=legend)

def plot_profile(taus, profiles, title, filename, colors, markers, logscale = False, legend = True):
    filename = os.path.join("plots", "Performance_profiles" , filename + ".pdf")
    sns.set(style="whitegrid", context="paper")
    plt.figure()
    fontsize = 20
    i = 0
    N = len(taus)
    for alg, prof in profiles.items():
        plt.plot(taus, prof, linewidth=2, label=alg, color = colors[i], marker = markers[i], markevery = int(N / 7)+1,markersize = 7, clip_on = False, zorder=3)
        i+= 1
    
    plt.xlim([1, max(taus)])
    plt.ylim([0, 1.05])
    plt.tick_params(axis="both", which="major", labelsize=fontsize - 2, bottom=True, left=True)
    plt.xlabel(r' $\tau$', fontsize= fontsize)
    plt.ylabel(r' $\rho(\tau)$', fontsize = fontsize)
    if logscale == True and np.max(taus) > 10:
        plt.xscale('log')
    plt.title(title, fontsize = fontsize)
    plt.grid(True)
    if legend: plt.legend(fontsize=fontsize-6)
    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    plt.close()

def plot_iter_obj(name):
    with open(f"{name}_allresults.pkl", "rb") as f:
        all_results = pickle.load(f)
    eps = 10**-20
    plt.figure()
    for size, result in all_results.items():
        algs = result.keys()
        colors = assign_color(algs)
        i = 0
        j = 0
        for alg, history in result.items():
            if j == 0:
                label = alg
            else:
                label = None
            plt.scatter(history['iteration'], np.array(history['obj']) + eps, label = label, color = colors[i])
            i += 1
        j += 1
    plt.yscale("log")
    plt.legend()
    plt.tight_layout()
    plt.title(f'Obj-Iter chart {size}')
    plt.show()


def create_report(results, name, position, parameters , name_workbook, with_fopt = True, avg_ite = None):
    if os.path.exists(name_workbook):
        workbook = load_workbook(name_workbook)
    else:
        workbook = Workbook()
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
    if name in workbook.sheetnames:
        worksheet = workbook[name]
    else:
        worksheet = workbook.create_sheet(name)

    row, col = position[0], position[1]

    if with_fopt:
        opt = find_min_value(results)
    else:
        opt = 0
    worksheet.cell(row=row, column=col, value=parameters["size"] + " " + parameters["additional_info"])
    row += 1
    worksheet.cell(row=row, column=col, value="data")
    worksheet.cell(row=row + 1, column=col, value=parameters["seed"])
    worksheet.cell(row=row, column=col + 1, value="k")
    worksheet.cell(row=row, column=col + 2, value="Residual")
    worksheet.cell(row=row, column=col + 3, value="Objective")
    worksheet.cell(row=row, column=col + 4, value="Time")
    if "lasso" in name: worksheet.cell(row=row, column=col + 5, value="Mse")        # Only report mse for the lasso problem
    row += 1

    for k, history_values in results.items():
        index = len(history_values["res"]) - 1
        if avg_ite != None: 
            worksheet.cell(row=row, column=col + 1, value=avg_ite[k])
        else: 
            worksheet.cell(row=row, column=col + 1, value=index + 1) 
        worksheet.cell(row=row, column=col + 2, value=history_values["res"][index])
        worksheet.cell(row=row, column=col + 3, value=history_values["obj"][index] - opt)
        worksheet.cell(row=row, column=col + 4, value=history_values["time"][index])
        if "lasso" in name: 
            worksheet.cell(row=row, column=col + 5, value=history_values["mse"][index])
            worksheet.cell(row=row, column=col + 6, value=k)
        else:
            worksheet.cell(row=row, column=col + 5, value=k)

        row += 1
    workbook.save(name_workbook)

def format_latex(results, name, with_fopt = True, avg_ite = None):
    output = []
    if with_fopt:
        opt = find_min_value(results)
    else:
        opt = 0
    excel_file_path = 'latex_support.xlsx'
    if not os.path.exists(excel_file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(excel_file_path)
    else:
        wb = openpyxl.load_workbook(excel_file_path)
        if name in wb.sheetnames:
            ws = wb[name]
        else:
            ws = wb.create_sheet(name)  

    row = ws.max_row + 1
    col = 1
    for k, history_values in results.items():
        index = len(history_values["res"]) - 1
        if avg_ite != None: 
            ws.cell(row=row, column=col , value=avg_ite[k])
        else: 
            ws.cell(row=row, column=col, value=index + 1) 
        ws.cell(row=row + 1, column=col, value=history_values["res"][index])
        ws.cell(row=row + 2, column=col, value=history_values["obj"][index] - opt)
        ws.cell(row=row + 3, column=col, value=history_values["time"][index])
        if "lasso" in name: 
            ws.cell(row=row+4, column=col, value=history_values["mse"][index])
        col += 1

    wb.save(excel_file_path)