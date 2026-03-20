import lasso, NMF, min_length_curve, dual_max_entropy, maximum_likelyhood, bcqp, bcfp
from save_and_plot import create_report, format_latex, plot_iter_obj, plot_performance_profile
from openpyxl import load_workbook
import numpy as np
import os
import itertools
import matplotlib.pyplot as plt
import pickle
name_workbook = 'Results_NPG.xlsx'

def clear_report_content(sheet):
    if os.path.exists(name_workbook):
        workbook = load_workbook(name_workbook)
        if sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.value = None
        workbook.save(name_workbook)

def find_min_value(results):
    f_min = np.inf
    for history_values in results.values():
        f_min = min(f_min, min(history_values["obj"]))
    return f_min

def prepare_report(size, row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results):
    avg_results = {}
    avg_iteration = {}
    if seed != seeds[0]:
        combined_results = all_results[size]
    else:
        combined_results = {}
    opt = find_min_value(results)
    for k, history_values in results.items():
        if k not in combined_results:
            combined_results[k] = {"res":[history_values["res"][-1]], "obj":[history_values["obj"][-1] - opt], "time":[history_values["time"][-1]],
                                        "iteration": [len(history_values["res"])], "steps": [1/np.mean(history_values['steps'])]}
            if "lasso" in name:
                combined_results[k]["mse"] = [history_values["mse"][-1]]

        else:
            combined_results[k]["res"].append(history_values["res"][-1])
            combined_results[k]["obj"].append(history_values["obj"][-1] - opt) 
            combined_results[k]["time"].append(history_values["time"][-1]) 
            combined_results[k]["iteration"].append(len(history_values["res"]))
            combined_results[k]["steps"].append(1/np.mean(history_values["steps"]))
            if "lasso" in name:
                combined_results[k]["mse"].append(history_values["mse"][-1])
        if seed == seeds[-1]:
            avg_iteration[k] = np.mean(combined_results[k]["iteration"])
            avg_results[k] = { "res":[np.mean(combined_results[k]["res"])], "obj":[np.mean(combined_results[k]["obj"])], 
                                "time":[np.mean(combined_results[k]["time"])] }
            if "lasso" in name: 
                avg_results[k]["mse"] = [np.mean(combined_results[k]["mse"])]
    if seed == seeds[0]:
        all_results[size] = combined_results
    create_report(results, name, (row, col), parameters, name_workbook)
    if seed == seeds[-1]:
        parameters["seed"] = "Average of all dataset."
        parameters["additional_info"] = ""
        create_report(avg_results, name, (row + len(results)+4, col), parameters, name_workbook, with_fopt=False, avg_ite=avg_iteration)
        format_latex(avg_results, name, with_fopt=False ,avg_ite=avg_iteration)
        

def experiment_lasso():
    clear_report_content("lasso")
    seeds = [1,2,3,4,5,6,7,8,9,10]
    size = [(n, k*n) for n in [512, 1024, 2048] for k in [2,4,8] if k*n <= 8192]
    all_results = {}
    row, col = 1,1
    for seed in seeds:
        for m, n in size:
            print(f"---------- Start seed = {seed}, m = {m}, n = {n} ----------")
            results, name_instance, name, parameters = lasso.run_lasso(m, n, seed)
            prepare_report((m, n), row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results)
            col += 8
        col = 1
        row += len(results) + 4
    with open(os.path.join("results", name + "_allresults.pkl"), "wb") as f:
        pickle.dump(all_results, f)
    plot_performance_profile(name = "lasso", legend=False)
    print("\n ================= Successfully run all experiments ! =================")

def experiment_min_len_curve():
    clear_report_content("min_len_curve")
    seeds = [1,2,3,4,5,6,7,8,9,10]
    size = [(50, 5000, 50000), (500, 5000, 1500), (2000, 5000, 200), (100, 10000, 50000), (1000, 10000, 1500), (2000, 10000, 500)]
    all_results = {}
    row, col = 1,1
    for seed in seeds:
        for m, n, N in size:
            print(f"---------- Start seed = {seed}, m = {m}, n = {n} ----------")
            results, name_instance, name, parameters = min_length_curve.run_min_len_curve(m, n, seed, N)
            prepare_report((m, n), row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results)
            col += 8
        col = 1
        row += len(results) + 4
    with open(os.path.join("results", name + "_allresults.pkl"), "wb") as f:
        pickle.dump(all_results, f)
    plot_performance_profile(name = "min_len_curve", legend=False)
    print("\n ================= Successfully run all experiments ! =================")



def experiment_maximum_likelyhood():
    clear_report_content("maximum_likelyhood")
    seeds = [1,2,3,4,5,6,7,8,9,10]
    size = [(100, 0.1, 10, 50, 3000), (100, 0.1, 10, 500, 200), (100, 0.1, 10, 1000, 100),(30, 0.1, 1000, 50, 5000), (50, 0.1, 1000, 100, 2500)]
    all_results = {}
    row, col = 1,1
    for seed in seeds:
        for n, lb, ub, M, N in size:
            print(f"---------- Start seed = {seed}, n = {n}, lb = {lb}, ub = {ub}, M = {M} ----------")
            results, name_instance, name, parameters = maximum_likelyhood.run_maximum_likelyhood(n, lb, ub, M, seed, N)
            prepare_report((n, lb, ub, M), row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results)
            col += 8
        col = 1
        row += len(results) + 4
    with open(os.path.join("results", name + "_allresults.pkl"), "wb") as f:
        pickle.dump(all_results, f)
    plot_performance_profile(name = "maximum_likelyhood", legend=False)
    print("\n ================= Successfully run all experiments ! =================")



def experiment_dual_max_entropy():
    clear_report_content("dual_max_ent")
    seeds = [1,2,3,4,5,6,7,8,9,10]
    size = [(100, 500, 100), (500, 2000, 100), (2000, 4000, 200), (4000, 5000, 200)]
    all_results = {}
    row, col = 1,1
    for seed in seeds:
        for m, n, N in size:
            print(f"---------- Start seed = {seed}, m = {m}, n = {n} ----------")
            results, name_instance, name, parameters = dual_max_entropy.run_dual_max_entropy(m, n, seed, N)
            prepare_report((m, n), row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results)
            col += 8
        col = 1
        row += len(results) + 4
    with open(os.path.join("results", name + "_allresults.pkl"), "wb") as f:
        pickle.dump(all_results, f)
    plot_performance_profile(name = "dual_max_ent", legend=False)
    print("\n ================= Successfully run all experiments ! =================")


def experiment_bcqp():
    clear_report_content("bcqp")
    seeds = [1,2,3,4,5,6,7,8,9,10]
    right_bound_Nmax = [(5,5000),(10, 10000)]
    size = [1000, 2000, 5000]
    all_results = {}
    row, col = 1,1
    for seed in seeds:
        for (r, N), n in list(itertools.product(right_bound_Nmax, size)):
            print(f"---------- Start seed = {seed}, r = {r}, n = {n}, maxIter = {N} ----------")
            results, name_instance, name, parameters = bcqp.run_bcqp(n,r, seed, N)
            prepare_report((n,r), row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results)
            col += 8
        col = 1
        row += len(results) + 4
    with open(os.path.join("results", name + "_allresults.pkl"), "wb") as f:
        pickle.dump(all_results, f)
    plot_performance_profile(name = "bcqp", legend=False)
    print("\n ================= Successfully run all experiments ! =================")


def experiment_bcfp():
    clear_report_content("bcfp")
    seeds = [1,2,3,4,5,6,7,8,9,10]
    right_bound = [5,10]
    size = [1000, 2000, 5000]
    all_results = {}
    row, col = 1,1
    for seed in seeds:
        for r, n in list(itertools.product(right_bound, size)):
            print(f"---------- Start seed = {seed}, r = {r}, n = {n} ----------")
            results, name_instance, name, parameters = bcfp.run_bcfp(n,r, seed)
            prepare_report((n,r), row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results)
            col += 8
        col = 1
        row += len(results) + 4
    with open(os.path.join("results", name + "_allresults.pkl"), "wb") as f:
        pickle.dump(all_results, f)
    plot_performance_profile(name = "bcfp", legend=False)
    print("\n ================= Successfully run all experiments ! =================")


def experiment_nmf():
    clear_report_content("nmf")
    seeds = [1,2,3,4,5,6,7,8,9,10]
    size = [(500, 20, 1000, 1000), (1000, 20, 500, 1000), (2000, 20, 3000, 1000), (3000, 20, 2000, 1000), (3000, 20, 3000, 1000), (500, 30, 1000, 1500), (1000, 30, 500, 1500), (2000, 30, 3000, 1500), (3000, 30, 2000, 1500), (3000, 30, 3000, 1500)]
    all_results = {}
    row, col = 1,1
    for seed in seeds:
        for m, r, n, N in size:
            print(f"---------- Start seed = {seed}, m = {m}, r = {r}, n = {n} ----------")
            results, name_instance, name, parameters = NMF.run_nmf(m, r, n, seed, N)
            name = "nmf"    # Rename the problem's name to have shorter name
            prepare_report((m, r, n), row, col, seed, seeds, results, name_instance, name, parameters, name_workbook, all_results)
            col += 8
        col = 1
        row += len(results) + 4
    with open(os.path.join("results", name + "_allresults.pkl"), "wb") as f:
        pickle.dump(all_results, f)
    plot_performance_profile(name = "nmf", legend=False)
    print("\n ================= Successfully run all experiments ! =================")





def main():
    experiment_lasso()
    # experiment_min_len_curve()
    # experiment_maximum_likelyhood()
    # experiment_dual_max_entropy()
    # experiment_bcqp()
    # experiment_bcfp()
    # experiment_nmf()


if __name__ == "__main__":
    main()

